from django.shortcuts import render, redirect

from django.views import generic
from django.urls import reverse_lazy

from django.contrib import messages

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
import json
import datetime

from datetime import timedelta
import calendar

from openpyxl import Workbook
import openpyxl
import locale 

# from django.utils import simplejson

# from .models import Client, Contract, Order


# from .forms import ClientForm, ContractForm

from sales.models import Contract, Order
from .models import Document, Report_day
from content.models import Video
from content.models import Playlist, Playlist_client_detail, Playlist_spot_detail, Playlist_document


def report_client(request):
    template_name = 'reportemision/contract_list.html'
    contexto = {}
    contract = Contract.objects.filter(state=True).all()

    # today = datetime.datetime.now()
    # order = Order.objects.filter(contract=id)
    if not contract:
        return HttpResponse('el registro con id: ' + str(id) + ' no existe!!!')

    if request.method == 'GET':

        contexto = {'obj': contract}

    return render(request, template_name, contexto)


def order_report_list(request, id):
    # id => contract_id
    template_name = 'reportemision/order_report_list.html'
    contexto = {}

    document = Document.objects.filter(contract=id).all()

    contexto = {'obj': document}
    return render(request, template_name, contexto)


def order_report_print(request, id):
    template_name = 'reportemision/order_report_print.html'
    contexto = {}

    document = Document.objects.get(pk=id)

    playlist_spot = Playlist_spot_detail.objects.filter(
        order_id=document.order.id, playlist_id=document.playlist_id).all()
    videos = []
    time_in_pauta_total = 0
    repeat_in_pauta = 0
    videos_id = []
    for item in playlist_spot:
        objeto = {}
        video = Video.objects.get(pk=item.video_id)
        videos_id.append(video.id)
        objeto['name_video'] = video.name
        objeto['duration'] = video.duration_all
        objeto['duration_seconds'] = video.duration_seconds
        objeto['repeat_in_pauta'] = item.repeat_count
        objeto['time_in_pauta'] = item.time_total
        time_in_pauta_total = time_in_pauta_total + int(item.time_total)
        repeat_in_pauta = repeat_in_pauta + int(item.repeat_count)
        
        videos.append(objeto)

    playlist = Playlist_document.objects.filter(
        playlist_id=document.playlist_id).all().order_by('order_list')
    playlist_client = []
    for item in playlist:
        objeto = {}
        objeto['orden_list'] = item.order_list
        if (not item.video.id in videos_id):
            objeto['name_video'] = 'Otro cliente'
            objeto['color'] = 0
        else:
            objeto['name_video'] = item.video.name
            objeto['color'] = 1
        objeto['duration'] = item.video.duration_all()
        playlist_client.append(objeto)
#
    # for item in playlist_client:
        # print(item)
    pass_day = int(repeat_in_pauta) * int(document.repeat_pauta_day)
    contexto = {'obj': document, 'videos': videos, 'time': time_in_pauta_total,
                'time2': repeat_in_pauta, 'playlist': playlist_client, 'pass_day':pass_day}

    return render(request, template_name, contexto)
# ============################################################


def report_client_all(request, id):
    template_name = 'reportemision/contract_list.html'
    contexto = {}

    contract = Contract.objects.get(pk=id)
    order = Order.objects.filter(contract=id)
    color = 1
    report = []
    # print(contract)
    for i in order:
        order_id = i.id
        # print(i.id)
        playlist_client_detail = Playlist_client_detail.objects.filter(
            order=order_id).order_by('playlist__create_date')
        if playlist_client_detail:
            for j in playlist_client_detail:

                # print(j.order.id)
                playlist_spot_detail = Playlist_spot_detail.objects.filter(
                    playlist=j.playlist.id, order=order_id)
                for k in playlist_spot_detail:
                    # print(k.video.name)
                    # se extrae informacion por videos de todos los contratos
                    item = {}
                    item["id"] = str(order_id) + str(k.video.id) + \
                        str(j.playlist.product.id)
                    item["contract_id"] = contract.id 
                    item["order_id"] = order_id
                    item["video_id"] = k.video.id
                    item["client"] = contract.client.name
                    item["start_date"] = contract.start_date
                    item["end_date"] = contract.end_date
                    item["playlist_id"] = j.playlist.id
                    item["location"] = j.playlist.product
                    item["create_date"] = j.playlist.create_date
                    item["video_name"] = k.video.name
                    item["video_duration"] = k.video.duration_all()
                    item["repeticiones"] = k.repeat_count
                    item["pauta_loop"] = j.pauta_loop
                    item["second_total"] = j.second_total
                    item["time_contract"] = j.time_contract
                    item["time_bonification"] = j.time_bonification
                    item["color"] = color
                    item["state"] = 1
                    item["proyection"] = k.playlist.proyection
                    # print(j.playlist.create_date)
                    report.append(item)
                if color == 1:
                    color = 0
                else:
                    color = 1

        else:
            # encaso de no existir informacion de los videos se toma todo en vacio
            item = {}
            item["id"] = str(order_id) + str(0) + str(i.product.id)
            item["contract_id"] = contract.id
            item["order_id"] = order_id
            item["video_id"] = 0
            item["client"] = contract.client.name
            item["start_date"] = contract.start_date
            item["end_date"] = contract.end_date
            item["playlist_id"] = 0
            item["state"] = 0
            item["location"] = i.product
            item["create_date"] = datetime.datetime.strptime(
                '0001-01-01', "%Y-%m-%d").date()
            item["video_name"] = 'sin registro de playlist'
            item["video_duration"] = 0
            item["repeticiones"] = 0
            item["pauta_loop"] = 0
            item["second_total"] = 0
            item["time_contract"] = 0
            item["time_bonification"] = 0
            item["color"] = 3
            item["proyection"] = False
            report.append(item)

        report_2 = report_resume(report)

    if request.method == 'GET':

        # for m in report:
        # print(m)

        template_name = 'reportemision/contract_date.html'
        contexto = {'report': report, 'contract': contract,
                    'order': order, 'report2': report_2}

        return render(request, template_name, contexto)

    if request.method == 'POST':

        bandera = request.POST.get('bandera')

        start_date_n = request.POST.get('start_date')
        end_date_n = request.POST.get('end_date')

        start_date_n = datetime.datetime.strptime(
            start_date_n, "%Y-%m-%d").date()
        end_date_n = datetime.datetime.strptime(end_date_n, "%Y-%m-%d").date()

        id_report = request.POST.getlist('id_report')
        print(id_report)

        if bandera == '1':
            aux_order = []
            print(bandera)
            for item in report_2:
                if item['id'] in id_report:
                    print(item['id'])
                    if (not item['order_id'] in aux_order):
                        aux_order.append(item['order_id'])
                        order_id = Order.objects.get(pk=item['order_id'])
                        contract_id = Contract.objects.get(pk=item['contract_id'])

                        if not item['playlist_id'] == 0:
                            playlist_id = Playlist.objects.get(
                                pk=item['playlist_id'])
                            document = Document(
                                start_date=start_date_n,
                                end_date=end_date_n,
                                contract=contract_id,
                                order=order_id,
                                repeat_pauta_day=item['pauta_loop'],
                                repeat_in_pauta=item['repeticiones'],

                                second_total_day=item['second_total'],
                                time_contract=item['time_contract'],
                                time_bonification=item['time_bonification'],
                                playlist=playlist_id,
                                user_created=request.user

                            )
                            document.save()
                        contract = Contract.objects.filter(state=True).all()

            template_name = 'reportemision/contract_list.html'

            contexto = {'obj': contract}
        # else:
        #     # diferencia = end_date_n - start_date_n
        #     # diferencia = diferencia.days + 1
        #     # print(diferencia)

        #     report_date = []
        #     # prueba para determinar fechas de inicio y final por cada lista creada
        #     randos_de_fechas = []
        #     aux = []
        #     aux_report = []
        #     for item in report:
        #         aux_report = report
        #         if not(item["id"] in aux):

        #             aux.append(item["id"])
        #             report_id = item["id"]
        #             randos_de_fechas = determinar_fechas(
        #                 aux_report, report_id, start_date_n, end_date_n)
        #             # fechas = rellenar_fechas(aux_report, report_id, start_date_n, end_date_n)
        #             # report_date.append(fechas)
        #             # print(report_date)

        #     template_name = 'reportemision/contract_complet.html'
        #     contexto = {'obj': report_date}
        #     # print(report_date)

        # return render(request, template_name, contexto)

    return render(request, template_name, contexto)


# para verificar clientes porfecha de creacion en cada lista de reprudoccion presente
def report_client_generate(request, id):
    template_name = 'reportemision/contract_list.html'
    contexto = {}

    contract = Contract.objects.get(pk=id)
    order = Order.objects.filter(contract=id)
    color = 1
    report = []
    # print(contract)
    for i in order:
        order_id = i.id
        # print(i.id)
        playlist_client_detail = Playlist_client_detail.objects.filter(
            order=order_id).order_by('playlist__create_date')
        if playlist_client_detail:
            for j in playlist_client_detail:

                playlist_spot_detail = Playlist_spot_detail.objects.filter(
                    playlist=j.playlist.id, order=order_id)
                for k in playlist_spot_detail:
                    date_start_report_day = Report_day.objects.filter(
                        order_id=order_id, video_id=k.video.id, contract_id=contract.id, playlist_id=j.playlist.id).order_by('date_now').first()

                    date_end_report_day = Report_day.objects.filter(
                        order_id=order_id, video_id=k.video.id, contract_id=contract.id, playlist_id=j.playlist.id).order_by('date_now').last()
                    
                    aux1 = ''
                    aux2 = ''

                    if not date_start_report_day:
                        aux1 = 'sin registros'
                    else:
                        aux1 = date_start_report_day.date_now
                    if not date_end_report_day:
                        aux2 = 'sin registros'
                    else:
                        aux2 = date_end_report_day.date_now
                        

                    count_report_day = Report_day.objects.filter(
                        order_id=order_id, video_id=k.video.id, contract_id=contract.id, playlist_id=j.playlist.id).count()
                    # print("contador: ")

                    # print(count_report_day)
                    # print(k.video.name)
                    # se extrae informacion por videos de todos los contratos
                    item = {}
                    item["id"] = str(order_id) + str(k.video.id) + \
                        str(j.playlist.product.id)
                    item["contract_id"] = contract.id
                    item["order_id"] = order_id
                    item["video_id"] = k.video.id
                    item["client"] = contract.client.name
                    item["start_date"] = contract.start_date
                    item["end_date"] = contract.end_date
                    item["playlist_id"] = j.playlist.id
                    item["location"] = j.playlist.product
                    item["create_date"] = j.playlist.create_date
                    item["video_name"] = k.video.name
                    item["video_duration"] = k.video.duration_all()
                    item["repeticiones"] = k.repeat_count
                    item["pauta_loop"] = j.pauta_loop
                    item["second_total"] = j.second_total
                    item["time_contract"] = j.time_contract
                    item["time_bonification"] = j.time_bonification
                    item["color"] = color
                    item["state"] = count_report_day

                    item["start_date_report"] = aux1
                    item["end_date_report"] = aux2
                    # print(j.playlist.create_date)
                    report.append(item)
                if color == 1:
                    color = 0
                else:
                    color = 1

        else:
            # encaso de no existir informacion de los videos se toma todo en vacio
            item = {}
            item["id"] = str(order_id) + str(0) + str(i.product.id)
            item["contract_id"] = contract.id
            item["order_id"] = order_id
            item["video_id"] = 0
            item["client"] = contract.client.name
            item["start_date"] = contract.start_date
            item["end_date"] = contract.end_date
            item["playlist_id"] = 0
            item["state"] = 'None'
            item["location"] = i.product
            item["create_date"] = datetime.datetime.strptime(
                '0001-01-01', "%Y-%m-%d").date()
            item["video_name"] = 'sin registro de playlist'
            item["video_duration"] = 0
            item["repeticiones"] = 0
            item["pauta_loop"] = 0
            item["second_total"] = 0
            item["time_contract"] = 0
            item["time_bonification"] = 0
            item["color"] = 3
            report.append(item)

    if request.method == 'GET':

        template_name = 'reportemision/contract_date_generate.html'
        contexto = {'report': report, 'contract': contract, 'order': order}

        return render(request, template_name, contexto)

    return render(request, template_name, contexto)

# ver todas las instalancias de report_day
def report_client_day_view(request, order_id, video_id, playlist_id): 
    template_name = 'reportemision/report_day_list.html'
    if request.method == 'GET':
        report_day = Report_day.objects.filter(order_id=order_id, video_id=video_id, playlist_id=playlist_id).order_by('date_now').all()

        print(order_id)
        print(video_id)
        print(playlist_id)
        contexto = {'obj':report_day}


    return render(request, template_name, contexto)

def report_day_delete(request, id):

    if request.method == 'GET':
        template_name = 'reportemision/report_day_delete.html'
        report_day = Report_day.objects.get(pk=id)
        contexto = {'obj':report_day}

    if request.method == 'POST':
        
        report_day = Report_day.objects.get(pk=id)
        report_day.delete()
        
        return HttpResponse('OK')

    return render(request,template_name, contexto)

#  utilizado por ajax para guardar las fechas individduales de uso en la base de datos
def report_generate_save(request):
    if request.method == 'POST':
        val = request.POST.get('val')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        days = end_date - start_date
        print(days.days)
        day_delta = datetime.timedelta(days=1)
        val = val.split('|')
        order_id = val[0]
        video_id = val[1]
        playlist_id = val[2]
        contract_id = val[3]

        order = Order.objects.get(pk=order_id)
        video = Video.objects.get(pk=video_id)
        playlist = Playlist.objects.get(pk=playlist_id)
        contract = Contract.objects.get(pk=contract_id)
        for x in range(days.days+1):
            report_day = Report_day(
                order=order,
                video=video,
                contract=contract,
                playlist=playlist,
                user_created=request.user,
                date_now=start_date + x*day_delta
            )
            report_day.save()

    return HttpResponse('OK')
# para ver la fechas en las que se generaron reportes


def report_client_view(request, id):
    template_name = 'reportemision/contract_date_view.html'
    contexto = {}

    if request.method == 'GET':
        contract = Contract.objects.get(pk=id)
        report_day = Report_day.objects.filter(contract_id=id).all()
        order = Order.objects.filter(contract_id=id).all()
        month_existing = []
        aux2 = []
        for item in report_day:
            objeto = {}
            date_now = item.date_now
            month = date_now.month
            year = date_now.year
            aux = str(item.order.id) + '-'+str(month) + '-' + str(year)
            report_month = str(month) + '-' + str(year)

            objeto['month_report'] = report_month
            objeto['order'] = item.order
            if not aux in month_existing:
                month_existing.append(aux)
                aux2.append(objeto)

        # print(month_existing)
        print(aux2)
        contexto = {'contract': contract, 'obj': aux2, 'order': order}

    return render(request, template_name, contexto)

# def separar lista por video()


def div_list(lista, video_id):
    nueva_lista = []

    for item in lista:
        if item.video.id == video_id:
            nueva_lista.append(item)

    return nueva_lista

# imprecion o ver el reporte ya por locaion y por fechas creo que es mejor descargar en xls


def report_client_xls(request, id, month, order_id):
    # locale.setlocale(locale.LC_ALL, 'es-bo')
    # Idioma "es-ES" (código para el español de España)
    if request.method == 'GET':
        # month_report = request.POST.get('month_report')
        # contract_id = request.POST.get('contract')
        # order_id = request.POST.get('order')
        # print(id)
        # print(month)
        month_s = month.split('-')[0]
        year = month.split('-')[1]
        contract = Contract.objects.get(pk=id)
        order = Order.objects.get(pk=order_id)
        report_day = Report_day.objects.filter(date_now__month=month_s, date_now__year=year,
                                               contract_id=id, order_id=order_id).all().order_by('order_id', 'video_id', 'date_now')

        video_aux = []
        new_report_day = []
        totales = []
        aux_date = []
        for item in report_day:

            if not item.video.id in video_aux:
                video_aux.append(item.video.id)
                aux_of_video = div_list(report_day, item.video.id)
                new_report_day.append(aux_of_video)

        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        # ws['B1'] = 'Sistema de Administración Publicitaria'
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'CLIENTE: '
        ws['C3'] = str(contract.client.name)
        ws['B4'] = 'FECHA DE INICIO: '
        ws['C4'] = str(contract.start_date)
        ws['B5'] = 'FECHA DE FINALIZACIÓN: '
        ws['C5'] = str(contract.end_date)

        ws['B7'] = 'PRODUCTO: '
        ws['C7'] = str(order.product)
        ws['B8'] = 'FUNCIONAMIENTO: '
        ws['C8'] = str(order.product.start_time) + ' a ' + str(order.product.end_time) + '. ' + str(order.product.time_operating_valid()) + ' fun'
        row = 10
        column = 5
        for item in new_report_day:
            aux_num = 31
            ws.merge_cells(start_row=int(row-1), start_column=int(column),
                           end_row=int(row-1), end_column=int(aux_num+column-1))
            for x in item:
                ws.cell(row=row-1, column=5).value = x.date_now.strftime('%B')
                spot_info = Playlist_spot_detail.objects.filter(
                    video_id=x.video.id, playlist=x.playlist.id).get()
                client_info = Playlist_client_detail.objects.filter( 
                    playlist=x.playlist.id, order_id=x.order.id).get()
                # print(client_info)
                column = 4
                day = x.date_now.day
                column = column + day
                ws.cell(row=row, column=column).value = x.date_now.day
                ws.cell(row=row, column=2).value = 'video: '
                ws.cell(row=row, column=3).value = str(x.video.name)
                ws.cell(row=(row+1), column=2).value = 'Duración: '
                ws.cell(row=(row+1), column=3).value = str(x.video.duration_all())
                ws.cell(row=(row+1), column=4).value = 'Pases dia: '
                ws.cell(row=(row+1), column=column).value = int(
                    client_info.pauta_loop) * int(spot_info.repeat_count)
                ws.cell(row=(row+2), column=4).value = 'Total (s): '

                ws.cell(row=(row+2), column=column).value = int(float(client_info.pauta_loop) * float(spot_info.repeat_count) * float(x.video.duration_seconds()))
                
                

                objeto = {}
                objeto['date_now'] = x.date_now
                objeto['client_info'] = client_info
                if not x.date_now in aux_date:
                    aux_date.append(x.date_now)
                    totales.append(objeto)

                
                
                # row = row + 1
                # column = column + 1
                # print(type(x.date_now))
                # print(x.date_now)

            row = row + 5
            column = 5

    aux_num = 31
    ws.merge_cells(start_row=int(row), start_column=int(column), end_row=int(row), end_column=int(aux_num+column-1))
    for item in totales:
        # print(item)
        column = 4
        day = item['date_now'].day
        column = column + day
        client_info_2 = item['client_info']
        ws.cell(row=(row+1), column=4).value = 'Pases dia: '
        ws.cell(row=(row+1), column=column).value = client_info_2.pauta_loop
        ws.cell(row=(row+2), column=4).value = 'Total (s): '
        ws.cell(row=(row+2), column=column).value = client_info_2.second_total
        ws.cell(row=(row+3), column=4).value = 'Contratado (s): '
        ws.cell(row=(row+3), column=column).value = client_info_2.time_contract
        ws.cell(row=(row+4), column=4).value = 'Bonificación (s): '
        ws.cell(row=(row+4), column=column).value = client_info_2.time_bonification
        
    for column_cells in ws.columns:
        new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
        new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length + 1
    
    
    today = datetime.datetime.now()

    # Establecemos el nombre del archivo
    nombre_archivo = "Reporte: " + str(today) + ".xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

   
# ======================================================================
def as_text(value):
    if value is None:
        return ""
    return str(value)

def report_resume(report):
    nueva_report = []
    aux_lista = []

    for item in report:
        if (not item["order_id"] in aux_lista):
            aux_lista.append(item["order_id"])
            # max_bonification = hallar_bonificacion_max(
            #     report, item["order_id"])
            if item["proyection"] == True:
                max_bonification = item
            else:
                max_bonification = hallar_bonificacion_max(report, item["order_id"])
            aux = extaer_playlist_id(report, max_bonification["playlist_id"])
            nueva_report = nueva_report + aux

    # for item in nueva_report:

    #     print(item)
    return(nueva_report)


def extaer_playlist_id(report, playlist_id):
    aux = []
    for item in report:
        if item["playlist_id"] == playlist_id:
            aux.append(item)

    return(aux)


def hallar_bonificacion_max(report, order_id):
    aux_max = 0
    for item2 in report:
        if item2["order_id"] == order_id:
            aux_max = item2
            pass

    for item in report:
        if item["order_id"] == order_id:
            if item["time_bonification"] >= aux_max["time_bonification"]:
                aux_max = item

    return(aux_max)
# ==================================================================
# prueba para determinar fechas de inicio y final por cada lista creada


# def determinar_fechas(lista, id, start_date_n, end_date_n):
#     aux_list = []
#     dias = end_date_n - start_date_n
#     dias = dias.days

#     for item in lista:
#         if item["id"] == id:
#             aux_list.append(item)
#     # print("---------")
#     # print(start_date_n)
#     # print(end_date_n)
#     # print('---------')

#     oficial = aux_list[0]

#     inicio = start_date_n
#     dia = start_date_n
#     lol = aux_list[0]
#     final = aux_list[len(aux_list)-1]["create_date"]
#     aux_final = 0  # datetime.datetime.strptime('', "%Y-%m-%d").date()
#     lista_final = []
#     if str(aux_list[0]["create_date"]) == '0001-01-01':
#         # print('no tiene registro de video ni de pauta')
#         pass
#     for item in aux_list:
#         if inicio > item['create_date']:
#             # aux_final = item['create_date']
#             pass
#         else:
#             if item['create_date'] > end_date_n:
#                 # aux_final = item['create_date']
#                 pass
#             else:
#                 day_play = item['create_date'] - inicio
#                 # print(type(day_play))
#                 # print(day_play)
#                 aux_final = item['create_date']
#                 # print('llego?')
#                 inicio = item['create_date']
#                 # item['start_date_pauta'] = inicio

#     if final > end_date_n:
#         # colocar un return
#         if aux_list[len(aux_list)-2]["create_date"]:
#             day_play = end_date_n - aux_list[len(aux_list)-2]["create_date"]
#             # print('lol')
#             # print(type(day_play))
#             # print(day_play)

#     else:
#         day_play = aux_final - end_date_n
#         # print(type(day_play))
#         # print(day_play)

#     pass


def rellenar_fechas(lista, id, start_date_n, end_date_n):

    aux_lista = []

    dias = end_date_n - start_date_n
    dias = dias.days

    for item in lista:
        if item["id"] == id:
            aux_lista.append(item)

    oficial = aux_lista[0]

    inicio = start_date_n
    dia = start_date_n
    lol = aux_lista[0]
    final = aux_lista[len(aux_lista)-1]["create_date"]
    lista_final = []
    # print(str(aux_lista[0]["create_date"]))
    if str(aux_lista[0]["create_date"]) == '0001-01-01':
        # print('no tiene registro de video ni de pauta')
        pass
    else:

        for k in aux_lista:
            aux = 0
            diferencia = k["create_date"] - inicio
            inicio = k["create_date"]
            # print(diferencia.days)

            if diferencia.days == 0:
                aux = 0
            if diferencia.days > 1:

                aux = diferencia.days
            if diferencia.days == 1:
                aux = 1

            # print(aux)

            for num in range(aux):
                # obj[str(dia)] = lol
                # lol["dia"] = dia
                a = {}
                a["dia"] = dia
                a["create_date"] = lol["create_date"]
                a["video_duration"] = lol["video_duration"]
                a["repeticiones"] = lol["repeticiones"]
                a["pauta_loop"] = lol["pauta_loop"]
                a["second_total"] = lol["second_total"]
                a["time_contract"] = lol["time_contract"]
                a["time_bonification"] = lol["time_bonification"]
                dia = dia + timedelta(days=1)
                lista_final.append(a)

            if dia == final:
                # obj[str(dia)] = k
                # k["dia"] = dia
                a = {}
                a["dia"] = dia
                a["create_date"] = k["create_date"]
                a["video_duration"] = k["video_duration"]
                a["repeticiones"] = k["repeticiones"]
                a["pauta_loop"] = k["pauta_loop"]
                a["second_total"] = k["second_total"]
                a["time_contract"] = k["time_contract"]
                a["time_bonification"] = k["time_bonification"]
                lista_final.append(a)

            lol = k

        if not(final == end_date_n):
            diferencia = end_date_n - final
            diferencia = diferencia.days + 1

            for num in range(diferencia):
                # obj[str(dia)] = aux_lista[len(aux_lista) - 1]
                lolo = aux_lista[len(aux_lista) - 1]
                # lolo["dia"] = dia
                a = {}
                a["dia"] = dia
                a["create_date"] = lolo["create_date"]
                a["video_duration"] = lol["video_duration"]
                a["repeticiones"] = lolo["repeticiones"]
                a["pauta_loop"] = lolo["pauta_loop"]
                a["second_total"] = lolo["second_total"]
                a["time_contract"] = lolo["time_contract"]
                a["time_bonification"] = lolo["time_bonification"]
                dia = dia + timedelta(days=1)
                lista_final.append(a)

    oficial["fechas"] = lista_final
  
    return oficial




# para 
