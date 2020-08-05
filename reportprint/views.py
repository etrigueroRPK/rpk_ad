from django.shortcuts import render, redirect

from django.views import generic
from django.urls import reverse_lazy

from django.contrib import messages

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
import json
import datetime

from openpyxl import Workbook

# from django.utils import simplejson

# from .models import Client, Contract, Order


# from .forms import ClientForm, ContractForm

from sales.models import Contract, Order
from content.models import Playlist, Playlist_client_detail, Playlist_spot_detail, Playlist_document


def report_contract(request, id):
    template_name = 'reportprint/report_contract.html'
    contexto = {}
    contract = Contract.objects.filter(pk=id).first()

    today = datetime.datetime.now()
    order = Order.objects.filter(contract=id)
    if not contract:
        return HttpResponse('el registro con id: ' + str(id) + ' no existe!!!')

    if request.method == 'GET':

        contexto = {'today': today, 'contract': contract, 'order': order}

    return render(request, template_name, contexto)


def report_playlist(request, id):
    template_name = 'reportprint/report_playlist.html'
    contexto = {}

    playlist = Playlist.objects.get(pk=id)
    playlist_id = playlist.id
    playlist_client = Playlist_client_detail.objects.filter(playlist=id).all()
    playlist_spot = Playlist_spot_detail.objects.filter(playlist=id).all()
    playlist_document = Playlist_document.objects.filter(playlist=id).all().order_by('order_list')
    # print(playlist_id)
    today = datetime.datetime.now()
    # order = Order.objects.filter(contract=id)

    if not playlist:
        return HttpResponse('el registro con id: ' + str(id) + ' no existe!!!')

    if request.method == 'GET':

        contexto = {'today': today, 'playlist': playlist, 'playlist_client': playlist_client,
                    'playlist_spot': playlist_spot, 'playlist_document': playlist_document}
        # print(today)

    return render(request, template_name, contexto)


def report_playlist_xls(request, id):

    playlist = Playlist.objects.get(pk=id)
    playlist_id = playlist.id
    playlist_client = Playlist_client_detail.objects.filter(playlist=id).all()
    playlist_spot = Playlist_spot_detail.objects.filter(playlist=id).all()
    playlist_document = Playlist_document.objects.filter(playlist=id).all()
    today = datetime.datetime.now()

    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'Sistema de Administración Publicitaria'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'LOCACION:'
    ws['C3'] = str(playlist.product)
    ws['B4'] = 'FECHA DE CREACION:'
    ws['C4'] = str(playlist.create_date)

    ws['B7'] = '#'
    ws['C7'] = 'Cliente, spot'
    ws['D7'] = 'Duración'
    ws['E7'] = 'Loop'
    ws['F7'] = 'tiempo total'
    ws['G7'] = 'porcentaje de spot'

    cont = 8
    cl = 1
    for item in playlist_spot:
        cli = str(item.video.contract.client) + ', ' + str(item.video.name)
        ws.cell(row=cont, column=2).value = cl
        ws.cell(row=cont, column=3).value = cli
        ws.cell(row=cont, column=4).value = item.video.duration_all()
        ws.cell(row=cont, column=5).value = item.repeat_count
        ws.cell(row=cont, column=6).value = item.time_total
        ws.cell(row=cont, column=7).value = item.porcentage

        cont = cont + 1
        cl = cl + 1
    cont = cont + 1
    ws.cell(row=cont, column=2).value = '#'
    ws.cell(row=cont, column=3).value = 'Cliente'
    ws.cell(row=cont, column=4).value = 'tiempo total'
    ws.cell(row=cont, column=5).value = 'loop de pauta'
    ws.cell(row=cont, column=6).value = 'tiempo (s)'
    ws.cell(row=cont, column=7).value = 'tiempo contratado'
    ws.cell(row=cont, column=8).value = 'tiempo de bonificación'
    cont = cont + 1

    cs = 1
    for item in playlist_client:
        cli = str(item.order.contract.client.name)
        ws.cell(row=cont, column=2).value = cs
        ws.cell(row=cont, column=3).value = cli
        ws.cell(row=cont, column=4).value = item.time_total
        ws.cell(row=cont, column=5).value = item.pauta_loop
        ws.cell(row=cont, column=6).value = item.second_total
        ws.cell(row=cont, column=7).value = item.time_contract
        ws.cell(row=cont, column=8).value = item.time_bonification

        cont = cont + 1
        cs = cs + 1
    cont = cont + 1
    ws.cell(row=cont, column=2).value = '#'
    ws.cell(row=cont, column=3).value = 'Video'
    cont = cont + 1

    csl = 1
    for item in playlist_document:
        cli = str(item.video.contract.client) +', '+ str(item.video)
        ws.cell(row=cont, column=2).value = csl
        ws.cell(row=cont, column=3).value = cli
        

        cont = cont + 1
        csl = csl + 1



    # Establecemos el nombre del archivo
    nombre_archivo = "Reporte: " + str(today) + ".xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
